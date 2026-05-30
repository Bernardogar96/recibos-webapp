import streamlit as st
import anthropic
import base64
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import tempfile
from PIL import Image
import io

# Mapeo de comercios
COMERCIOS_CAMPOS = {
    "carnes_finas_san_juan": {
        "nombre": "Carnes Finas San Juan",
        "campos": ["# Transacción", "Fecha", "Monto"]
    },
    "iconn_petro7": {
        "nombre": "ICONN (Petro 7)",
        "campos": ["# Estación", "Folio", "WebID", "Monto"]
    },
    "sams_club": {
        "nombre": "SAM'S Club",
        "campos": ["TC", "TR", "Fecha", "Monto", "IVA"]
    },
    "walmart": {
        "nombre": "Walmart",
        "campos": ["TC", "TR", "Fecha", "Monto", "IVA"]
    },
    "farmacias_ahorro": {
        "nombre": "Farmacias del Ahorro",
        "campos": ["ITU", "Fecha", "Monto"]
    },
    "heb": {
        "nombre": "HEB",
        "campos": ["Sucursal", "6 Dígitos bajo fecha", "Fecha", "Monto"]
    },
    "home_depot": {
        "nombre": "Home Depot",
        "campos": ["# Ticket (bajo código barras)", "Fecha", "Monto"]
    },
    "angelly": {
        "nombre": "Angelly",
        "campos": ["Código de Facturación", "Fecha", "Monto"]
    },
    "oxxo_gas": {
        "nombre": "OXXO GAS",
        "campos": ["Estación", "Folio", "Cantidad", "Monto"]
    },
    "petromax": {
        "nombre": "PETROMAX",
        "campos": ["Estación", "Folio", "WebID", "Monto"]
    }
}

def encode_image_to_base64(image_bytes):
    """Convierte bytes de imagen a base64"""
    return base64.standard_b64encode(image_bytes).decode('utf-8')

def convertir_a_jpg(image_bytes, filename):
    """
    Detecta formato y convierte a JPG si es necesario
    Retorna: (jpg_bytes, formato_original, fue_convertida)
    Maneja: JPG, PNG, HEIC, WEBP, BMP, TIFF, GIF, MPO
    """
    try:
        # Detectar formato por extensión o contenido
        ext = filename.split('.')[-1].upper()
        
        # Para JPG/JPEG, intentar primero sin conversión
        if ext in ['JPEG', 'JPG']:
            try:
                img = Image.open(io.BytesIO(image_bytes))
                if img.format and img.format.upper() in ['JPEG', 'JPG']:
                    return image_bytes, 'JPG', False
            except Exception:
                pass  # Seguir con conversión
        
        # Intentar abrir con PIL
        img = Image.open(io.BytesIO(image_bytes))
        original_format = img.format or ext
        
        # Si ya es JPG válido, retorna tal cual
        if original_format.upper() in ['JPEG', 'JPG']:
            return image_bytes, original_format, False
        
        # Convertir a RGB si es necesario
        if img.mode in ('RGBA', 'LA', 'P'):
            rgb_img = Image.new('RGB', img.size, (255, 255, 255))
            if img.mode == 'RGBA':
                rgb_img.paste(img, mask=img.split()[3])
            else:
                rgb_img.paste(img)
            img = rgb_img
        elif img.mode != 'RGB':
            img = img.convert('RGB')
        
        # Guardar como JPG en memoria con calidad alta
        jpg_buffer = io.BytesIO()
        img.save(jpg_buffer, format='JPEG', quality=95, optimize=False)
        jpg_bytes = jpg_buffer.getvalue()
        
        return jpg_bytes, original_format, True
    
    except Exception as e:
        error_msg = str(e)[:50]
        st.warning(f"⚠️ {filename}: {error_msg}")
        return None, filename.split('.')[-1].upper(), False

def procesar_recibo_vision(client, image_base64, imagen_nombre):
    """Extrae datos del recibo con Claude Vision"""
    
    comercios_info = "\n".join([
        f"- {datos['nombre']}: campos = {', '.join(datos['campos'])}"
        for datos in COMERCIOS_CAMPOS.values()
    ])
    
    try:
        message = client.messages.create(
            model="claude-opus-4-6",
            max_tokens=1024,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "image",
                            "source": {
                                "type": "base64",
                                "media_type": "image/jpeg",
                                "data": image_base64
                            }
                        },
                        {
                            "type": "text",
                            "text": f"""Analiza este recibo y extrae los datos.

COMERCIOS Y CAMPOS REQUERIDOS:
{comercios_info}

Responde SOLO con JSON (sin markdown):
{{
  "comercio": "nombre exacto del comercio detectado",
  "campos_extraidos": {{
    "campo1": "valor",
    "campo2": "valor"
  }},
  "monto_total": número,
  "fecha": "YYYY-MM-DD"
}}

IMPORTANTE:
- Si no encuentras un campo, pon null
- Montos sin símbolo: 1234.56
- Usa nombres de campos EXACTOS"""
                        }
                    ]
                }
            ]
        )
        
        respuesta = message.content[0].text.strip()
        if '```' in respuesta:
            respuesta = respuesta.split('```')[1]
            if respuesta.startswith('json'):
                respuesta = respuesta[4:]
            respuesta = respuesta.split('```')[0]
        
        datos = json.loads(respuesta)
        return datos
    except Exception as e:
        return None

def crear_excel_dinamico(datos_recibos):
    """Crea Excel con columnas específicas por comercio"""
    
    wb = Workbook()
    recibos_por_comercio = {}
    
    for datos in datos_recibos:
        comercio = datos.get('comercio', 'Desconocido')
        if comercio not in recibos_por_comercio:
            recibos_por_comercio[comercio] = []
        recibos_por_comercio[comercio].append(datos)
    
    wb.remove(wb.active)
    
    for comercio_key, recibos in recibos_por_comercio.items():
        campos_info = None
        for key, info in COMERCIOS_CAMPOS.items():
            if info['nombre'].lower() == comercio_key.lower():
                campos_info = info
                break
        
        if not campos_info:
            campos_info = {'nombre': comercio_key, 'campos': list(recibos[0].get('campos_extraidos', {}).keys())}
        
        sheet = wb.create_sheet(comercio_key[:31])
        headers = ['#'] + campos_info['campos'] + ['Archivo', 'Notas']
        sheet.append(headers)
        
        header_fill = PatternFill(start_color='0A4980', end_color='0A4980', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for i, datos in enumerate(recibos, 1):
            campos_ext = datos.get('campos_extraidos', {})
            row = [i]
            
            for campo in campos_info['campos']:
                valor = campos_ext.get(campo, '')
                row.append(valor)
            
            row.append(datos.get('archivo', ''))
            row.append('')
            
            sheet.append(row)
            
            for cell in sheet[sheet.max_row]:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        for i, header in enumerate(headers, 1):
            sheet.column_dimensions[chr(64+i)].width = max(15, len(str(header)))
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ============ STREAMLIT APP ============

st.set_page_config(page_title="📸 Automatizador de Recibos", layout="wide")

st.title("📸 Automatizador de Recibos")
st.markdown("**Sube fotos de recibos → Obten Excel con datos extraídos automáticamente**")

# Verificar API key
api_key = st.secrets.get("ANTHROPIC_API_KEY") if hasattr(st, 'secrets') else None

if not api_key:
    st.error("⚠️ Falta configurar ANTHROPIC_API_KEY en los secrets de Streamlit")
    st.info("Para desplegar: crea un `secrets.toml` con tu clave de Anthropic")
    st.stop()

client = anthropic.Anthropic(api_key=api_key)

# Sidebar - Instrucciones
with st.sidebar:
    st.markdown("### 📋 Instrucciones")
    st.markdown("""
    1. **Sube imágenes** de recibos (JPG, PNG, HEIC)
    2. **Selecciona comercios** que quieres procesar
    3. **Haz clic en "Procesar"**
    4. **Descarga** el Excel con datos extraídos
    
    ### ✅ Comercios soportados
    - SAM'S Club (TC, TR)
    - Walmart (TC, TR)
    - HEB (Sucursal, 6 dígitos)
    - OXXO GAS (Estación, Folio)
    - PETROMAX (Estación, Folio, WebID)
    - ICONN Petro 7
    - Home Depot
    - Farmacias del Ahorro
    - Y más...
    """)

# Layout principal
col1, col2 = st.columns([2, 1])

with col1:
    st.subheader("📤 Sube tus recibos")
    uploaded_files = st.file_uploader(
        "Selecciona imágenes",
        type=["jpg", "jpeg", "png", "heic", "webp", "bmp", "tiff", "gif"],
        accept_multiple_files=True,
        help="📸 Cualquier formato: HEIC, PNG, WEBP, BMP, TIFF, GIF, JPG. Se convierten automáticamente a JPG"
    )
    
    if uploaded_files:
        st.success(f"✅ {len(uploaded_files)} imagen(es) cargada(s)")
        
        # Preview
        with st.expander("👀 Ver preview de imágenes"):
            cols = st.columns(min(3, len(uploaded_files)))
            for idx, file in enumerate(uploaded_files[:6]):
                with cols[idx % 3]:
                    st.image(file, caption=file.name, width=150)
                    
        # Procesar
        if st.button("🚀 Procesar Recibos", type="primary", use_container_width=True):
            progress_bar = st.progress(0)
            status_text = st.empty()
            results_container = st.container()
            
            datos_recibos = []
            errores = []
            
            for idx, uploaded_file in enumerate(uploaded_files):
                # Actualizar progreso
                progress = (idx + 1) / len(uploaded_files)
                progress_bar.progress(progress)
                
                try:
                    # Leer archivo original
                    image_bytes = uploaded_file.read()
                    
                    # Convertir a JPG si es necesario
                    jpg_bytes, formato_original, fue_convertida = convertir_a_jpg(image_bytes, uploaded_file.name)
                    
                    if jpg_bytes is None:
                        errores.append(f"{uploaded_file.name}: No se pudo convertir")
                        continue
                    
                    # Mostrar estado de conversión
                    conversion_text = f" ({formato_original}→JPG)" if fue_convertida else " (JPG nativo)"
                    status_text.text(f"Procesando {idx + 1}/{len(uploaded_files)}: {uploaded_file.name}{conversion_text}")
                    
                    # Codificar a base64
                    image_base64 = encode_image_to_base64(jpg_bytes)
                    
                    # Procesar con Vision
                    datos = procesar_recibo_vision(client, image_base64, uploaded_file.name)
                    
                    if datos:
                        datos['archivo'] = uploaded_file.name
                        datos['formato_original'] = formato_original
                        datos_recibos.append(datos)
                    else:
                        errores.append(uploaded_file.name)
                
                except Exception as e:
                    errores.append(f"{uploaded_file.name}: {str(e)[:50]}")
            
            # Resultados
            st.success("✅ Procesamiento completado")
            
            col_success, col_error = st.columns(2)
            with col_success:
                st.metric("✅ Exitosos", len(datos_recibos))
            with col_error:
                if errores:
                    st.metric("❌ Errores", len(errores))
            
            # Mostrar errores si los hay
            if errores:
                with st.expander("⚠️ Imágenes con error"):
                    for error in errores:
                        st.caption(f"❌ {error}")
            
            # Generar Excel
            if datos_recibos:
                st.subheader("📊 Datos Extraídos")
                
                # Resumen por comercio
                comercios_count = {}
                for datos in datos_recibos:
                    comercio = datos.get('comercio', 'Desconocido')
                    comercios_count[comercio] = comercios_count.get(comercio, 0) + 1
                
                cols = st.columns(len(comercios_count))
                for idx, (comercio, count) in enumerate(comercios_count.items()):
                    with cols[idx % len(cols)]:
                        st.metric(comercio, count)
                
                # Crear Excel
                excel_bytes = crear_excel_dinamico(datos_recibos)
                
                # Descargar
                st.download_button(
                    label="⬇️ Descargar Excel",
                    data=excel_bytes,
                    file_name="recibos_extraidos.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary"
                )
                
                # Mostrar preview de datos
                with st.expander("📋 Vista previa de datos"):
                    for comercio, recibos in comercios_count.items():
                        st.write(f"**{comercio}** ({recibos} recibos)")
                        for i, datos in enumerate([d for d in datos_recibos if d.get('comercio') == comercio][:3], 1):
                            with st.container():
                                cols = st.columns([1, 3])
                                with cols[0]:
                                    st.caption(f"#{i}")
                                with cols[1]:
                                    campos = datos.get('campos_extraidos', {})
                                    for campo, valor in campos.items():
                                        st.caption(f"**{campo}:** {valor}")
                        st.divider()

# Sidebar derecho - Info
with col2:
    st.subheader("ℹ️ Información")
    st.info("""
    ### ⚡ Características
    - ✅ Procesa múltiples imágenes
    - ✅ Identifica comercio automáticamente
    - ✅ Extrae campos específicos
    - ✅ Genera Excel organizado
    - ✅ Descarga instantánea
    
    ### 💰 Costo
    ~$0.02 USD por recibo
    
    ### 🔒 Privacidad
    Las imágenes se procesan con Claude Vision (Anthropic)
    No se guardan en el servidor
    """)
